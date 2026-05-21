import os
import cv2
import face_recognition
import pandas as pd
from datetime import datetime
from tkinter import *
from tkinter import messagebox, filedialog
from fpdf import FPDF
from openpyxl import Workbook

# Constants
KNOWN_FACES_DIR = "known_faces"
STUDENT_CSV = "students.csv"
ATTENDANCE_FILE = "attendance.xlsx"
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin123"

# Ensure known_faces directory exists
os.makedirs(KNOWN_FACES_DIR, exist_ok=True)

# Create attendance file with headers if it doesn't exist
if not os.path.exists(ATTENDANCE_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["ID", "Registration Number", "Name", "Roll Number", "Date", "Image"])
    wb.save(ATTENDANCE_FILE)

# Load student data
def load_students():
    if os.path.exists(STUDENT_CSV):
        try:
            return pd.read_csv(STUDENT_CSV)
        except:
            return pd.DataFrame(columns=["ID", "RegNo", "Name", "RollNumber", "Image"])
    return pd.DataFrame(columns=["ID", "RegNo", "Name", "RollNumber", "Image"])

students_df = load_students()

# Load known face encodings
known_encodings = []
known_names = []

print("Loading known face encodings...")
for _, row in students_df.iterrows():
    image_path = row["Image"]
    if os.path.exists(image_path):
        image = face_recognition.load_image_file(image_path)
        encodings = face_recognition.face_encodings(image)
        if encodings:
            known_encodings.append(encodings[0])
            known_names.append(row["Name"])

# Mark attendance
def mark_attendance(name):
    student = students_df[students_df["Name"].str.lower().str.strip() == name.lower().strip()]
    if student.empty:
        return

    id_ = student.iloc[0]["ID"]
    reg = student.iloc[0]["RegNo"]
    roll = student.iloc[0]["RollNumber"]
    img = student.iloc[0]["Image"]
    date_str = datetime.now().strftime("%d-%m-%Y")

    if os.path.exists(ATTENDANCE_FILE):
        df = pd.read_excel(ATTENDANCE_FILE, engine='openpyxl')
    else:
        df = pd.DataFrame(columns=["ID", "Registration Number", "Name", "Roll Number", "Date", "Image"])

    if not ((df["Name"] == name) & (df["Date"] == date_str)).any():
        new_entry = pd.DataFrame([{
            "ID": id_,
            "Registration Number": reg,
            "Name": name,
            "Roll Number": roll,
            "Date": date_str,
            "Image": img
        }])
        df = pd.concat([df, new_entry], ignore_index=True)
        df.to_excel(ATTENDANCE_FILE, index=False, engine='openpyxl')
        print(f"Attendance marked for {name}")

# Facial recognition with camera preview
def run_recognition():
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        messagebox.showerror("Error", "Camera not accessible.")
        return

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        locations = face_recognition.face_locations(rgb)
        encodings = face_recognition.face_encodings(rgb, locations)

        for encoding, location in zip(encodings, locations):
            matches = face_recognition.compare_faces(known_encodings, encoding)
            distances = face_recognition.face_distance(known_encodings, encoding)
            name = "Unknown"

            if True in matches:
                best_index = distances.argmin()
                name = known_names[best_index]
                mark_attendance(name)

                top, right, bottom, left = location
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
                cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

        cv2.imshow("Attendance System - Press Q to Exit", frame)
        if cv2.waitKey(1) & 0xFF == ord("q"):
            break

    cap.release()
    cv2.destroyAllWindows()

# Register new student
def register_student():
    def save_student():
        name = name_entry.get().strip()
        roll = roll_entry.get().strip()
        reg = reg_entry.get().strip()

        if not all([name, roll, reg]):
            messagebox.showerror("Error", "All fields are required.")
            return

        image_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg *.png *.jpeg")])
        if not image_path:
            return

        image = face_recognition.load_image_file(image_path)
        encodings = face_recognition.face_encodings(image)
        if not encodings:
            messagebox.showerror("Error", "No face found in image.")
            return

        filename = os.path.join(KNOWN_FACES_DIR, f"{name}.jpg")
        cv2.imwrite(filename, cv2.imread(image_path))

        new_id = len(students_df) + 1
        students_df.loc[len(students_df)] = [new_id, reg, name, roll, filename]
        students_df.to_csv(STUDENT_CSV, index=False)
        messagebox.showinfo("Success", "Student registered.")
        reg_window.destroy()

    reg_window = Toplevel(root)
    reg_window.title("Register Student")
    Label(reg_window, text="Name").grid(row=0, column=0)
    name_entry = Entry(reg_window)
    name_entry.grid(row=0, column=1)

    Label(reg_window, text="Roll Number").grid(row=1, column=0)
    roll_entry = Entry(reg_window)
    roll_entry.grid(row=1, column=1)

    Label(reg_window, text="Registration Number").grid(row=2, column=0)
    reg_entry = Entry(reg_window)
    reg_entry.grid(row=2, column=1)

    Button(reg_window, text="Save", command=save_student).grid(row=3, columnspan=2, pady=10)

# View attendance
def view_attendance():
    date = date_entry.get().strip()
    try:
        df = pd.read_excel(ATTENDANCE_FILE, engine='openpyxl')
        if date:
            df = df[df["Date"] == date]
        report_text.delete("1.0", END)
        report_text.insert(END, df.to_string(index=False) if not df.empty else "No records found.")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# Export to PDF
def export_pdf():
    date = date_entry.get().strip()
    try:
        df = pd.read_excel(ATTENDANCE_FILE, engine='openpyxl')
        if date:
            df = df[df["Date"] == date]
        if df.empty:
            messagebox.showinfo("Info", "No records to export.")
            return

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt="Attendance Report", ln=True, align='C')

        for _, row in df.iterrows():
            line = f"{row['ID']} | {row['Registration Number']} | {row['Name']} | {row['Roll Number']} | {row['Date']}"
            pdf.cell(200, 10, txt=line, ln=True)

        path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
        if path:
            pdf.output(path)
            messagebox.showinfo("Success", "PDF exported successfully.")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# Admin login
def admin_login():
    def verify():
        if user_entry.get() == ADMIN_USERNAME and pass_entry.get() == ADMIN_PASSWORD:
            login_window.destroy()
            main_window()
        else:
            messagebox.showerror("Error", "Incorrect credentials.")

    login_window = Tk()
    login_window.title("Admin Login")
    login_window.geometry("300x150")
    Label(login_window, text="Username").pack()
    user_entry = Entry(login_window)
    user_entry.pack()
    Label(login_window, text="Password").pack()
    pass_entry = Entry(login_window, show="*")
    pass_entry.pack()
    Button(login_window, text="Login", command=verify).pack(pady=10)
    login_window.mainloop()

# GUI
def main_window():
    global root, date_entry, report_text
    root = Tk()
    root.title("Face Recognition Attendance System")
    root.geometry("700x600")

    Label(root, text="Facial Recognition Attendance System", font=("Helvetica", 16, "bold")).pack(pady=10)
    Button(root, text="Start Attendance", command=run_recognition, width=25, bg="green", fg="white").pack(pady=5)
    Button(root, text="Register Student", command=register_student, width=25, bg="blue", fg="white").pack(pady=5)

    frame = Frame(root)
    frame.pack(pady=10)

    Label(frame, text="Date (DD-MM-YYYY):").pack(side=LEFT)
    date_entry = Entry(frame)
    date_entry.pack(side=LEFT, padx=5)

    Button(frame, text="View Attendance", command=view_attendance).pack(side=LEFT, padx=5)
    Button(frame, text="Export PDF", command=export_pdf).pack(side=LEFT, padx=5)

    report_text = Text(root, height=20, width=80)
    report_text.pack(pady=10)

    root.mainloop()

# Start with admin login
admin_login()
